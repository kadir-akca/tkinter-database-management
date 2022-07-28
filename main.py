import tkinter as tk
import xml.dom.minidom
from datetime import datetime
from tkinter import LEFT, ttk, END

import openpyxl
import pyperclip

import database
import fonts
import methods
from database import open_db
from methods import generate_id_random

x = database.get_last_experienceid()


class PageClass(tk.Frame):
    def __init__(self, *args, **kwargs):
        tk.Frame.__init__(self, *args, **kwargs)

    def show(self):
        self.lift()


class DatabaseManagement(tk.Frame):
    def __init__(self, *args, **kwargs):
        tk.Frame.__init__(self, *args, **kwargs)

        p1 = PageClassOne(self)
        p2 = PageClassTwo(self)

        buttonframe = tk.Frame(self)
        container = tk.Frame(self)
        buttonframe.pack(side="bottom", fill="x", expand=False)
        container.pack(side="top", fill="both", expand=True)

        p1.place(in_=container, x=0, y=0, relwidth=1, relheight=1)
        p2.place(in_=container, x=0, y=0, relwidth=1, relheight=1)

        b1 = tk.Button(buttonframe, text='First Page', command=p1.show)
        b2 = tk.Button(buttonframe, text='Second Page', command=p2.show)

        b_opendb = tk.Button(buttonframe, text='Open Database', command=lambda: open_db())
        b_quit = tk.Button(buttonframe, text='Exit', command=quit)

        b1.pack(side=LEFT)
        b2.pack(side=LEFT)
        # b3.pack(side=LEFT)
        b_opendb.pack(side=LEFT)
        b_quit.pack(side=LEFT)
        # b_reset.pack(side=LEFT)
        p1.show()


class PageClassOne(PageClass):

    def __init__(self, *args, **kwargs):
        PageClass.__init__(self, *args, **kwargs)
        self.confirm_l = None
        self.username_e = None
        self.conn = None
        self.x = None
        self.date = None
        self.part_name = None
        self.part_number = None
        self.inspector_name = None
        self.department_name = None
        self.experience_id = None
        self.product_name = None
        self.val_dep = ['HQ', 'Stuttgart', 'HongKong', 'China']
        self.val_exptype = ['BB150', 'BB300', 'BA300', 'BB500']
        self.val_artsn = [['BB150xxx', 'BB150yyy', 'BB150zzz'],
                          ['BB300xxx', 'BB300yyy', 'BB300zzz'],
                          ['BA300xxx', 'BA300yyy', 'BA300zzz'],
                          ['BB500xxx', 'BB500yyy', 'BB500zzz']]
        self.filename = ''

        label = tk.Label(self, text='New Test Page', font=fonts.LARGE_FONT)
        label.grid(row=0, column=0, pady=10, padx=10, sticky="NSEW", columnspan=3)

        self.deviceSN = tk.Entry(self)
        self.deviceSN.grid(row=1, column=1, columnspan=2, padx=5, pady=5, ipadx=5, ipady=2, sticky="NSEW")

        self.operator = tk.StringVar()
        self.cbb_operator = ttk.Combobox(self, textvariable=self.operator)
        self.cbb_operator.grid(row=2, column=1, columnspan=2, padx=5, pady=5, ipadx=5, ipady=2, sticky="NSEW")
        self.cbb_operator['values'] = database.get_operators()
        self.cbb_operator['state'] = 'readonly'

        self.department = tk.StringVar()
        self.cbb_department = ttk.Combobox(self, textvariable=self.department)
        self.cbb_department.grid(row=3, column=1, columnspan=2, padx=5, pady=5, ipadx=5, ipady=2, sticky="NSEW")
        self.cbb_department['values'] = self.val_dep
        self.cbb_department['state'] = 'readonly'

        self.experiencetype = tk.StringVar()
        self.cbb_experience_type = ttk.Combobox(self, textvariable=self.experiencetype)
        self.cbb_experience_type.bind('<<ComboboxSelected>>', self.list_artefactSN)
        self.cbb_experience_type.grid(row=4, column=1, columnspan=2, padx=5, pady=5, ipadx=5, ipady=2, sticky="NSEW")
        self.cbb_experience_type['values'] = self.val_exptype
        self.cbb_experience_type['state'] = 'readonly'

        self.artefactSN = tk.StringVar()
        self.artefactSN_cbb = ttk.Combobox(self, textvariable=self.artefactSN)
        self.artefactSN_cbb.grid(row=5, column=1, columnspan=2, padx=5, pady=5, ipadx=5, ipady=2, sticky="NSEW")
        self.artefactSN_cbb['state'] = 'readonly'

        self.certificateSN = tk.Entry(self)
        self.certificateSN.insert(0, 'certificate sn will be shown here automatically')
        self.certificateSN.grid(row=6, column=1, padx=5, columnspan=2, pady=5, ipadx=5, ipady=2, sticky="NSEW")
        self.certificateSN.config(state='disabled')

        self.subject = tk.Text(self, width=20, height=3)
        self.subject.grid(row=7, column=1, columnspan=2, padx=5, pady=20, ipadx=5, ipady=2, sticky="NSEW")

        self.experienceid = tk.Entry(self)
        self.experienceid.grid(row=9, column=1, padx=5, pady=5, ipadx=5, ipady=2, sticky="NSEW")

        deviceSN_l = tk.Label(self, text="Device SN:")
        deviceSN_l.grid(row=1, column=0)
        operator_l = tk.Label(self, text="Operator")
        operator_l.grid(row=2, column=0)
        department_l = tk.Label(self, text="Department:")
        department_l.grid(row=3, column=0)
        experiencetype = tk.Label(self, text="Experience Type:")
        experiencetype.grid(row=4, column=0)
        artefactSN_l = tk.Label(self, text="Artefact SN:")
        artefactSN_l.grid(row=5, column=0)
        certificateSN_l = tk.Label(self, text="Certificate SN:")
        certificateSN_l.grid(row=6, column=0)
        subject_l = tk.Label(self, text="Subject:")
        subject_l.grid(row=7, column=0)
        experienceid_l = tk.Label(self, text="Experience ID:")
        experienceid_l.grid(row=9, column=0)

        self.button1 = tk.Button(self, text='Fill with XML', command=lambda: self.get_data_from_xml())
        self.button1.grid(row=8, column=0, columnspan=3, sticky="NSEW")

        self.button2 = tk.Button(self, text='Generate Experience ID', command=lambda: self.generate_exp_id_random())
        self.button2.grid(row=9, column=2, padx=5, pady=5, sticky="NSEW")

        self.button3 = tk.Button(self, text='Create Experience', command=lambda: self.experience_to_db())
        self.button3.grid(row=10, column=0, columnspan=3, padx=5, pady=5, sticky="NSEW")

        button3 = tk.Button(self, text='Add operator',
                            command=lambda: self.add_user())
        button3.grid(row=2, column=3, sticky="NSEW")

        '''line1 = tk.Label(self, text='_______________________________________________________')
        line1.grid(row=10, column=0, columnspan=3)

        experienceid_ll = tk.Label(self, text="Experience ID:")
        experienceid_ll.grid(row=11, column=0)

        self.experienceid2 = tk.Entry(self)
        self.experienceid2.grid(row=11, column=1, padx=5, pady=5, ipadx=5, ipady=2, sticky="NSEW")

        copy = tk.Button(self, command=lambda: self.copy_to_clipboard(), text='Copy to Clipboard')
        copy.grid(row=11, column=2, padx=5, pady=5, sticky="NSEW")

        self.copy_l = tk.Label(self)
        self.copy_l.grid(row=11, column=3)'''

        self.b_reset = tk.Button(self, command=lambda: self.reset(), text='Reset all')
        self.b_reset.grid(row=12, column=0, columnspan=3, padx=5, pady=5, sticky="NSEW")

    def experience_to_db(self):
        dev = self.deviceSN.get()
        op = self.operator.get()
        dep = self.department.get()
        etype = self.experiencetype.get()
        artsn = self.artefactSN.get()
        certsn = self.certificateSN.get()
        sub = self.subject.get('1.0', END)
        eid = self.experienceid.get()
        database.insert_accuracy_test(dev, op, dep, etype, artsn, certsn, sub, eid)

    def add_user(self):
        adduser_window = tk.Toplevel(root)
        adduser_window.title('Add new operator')
        adduser_window.geometry('200x200')
        username_l = tk.Label(adduser_window, text='Username')
        username_l.grid(row=0, column=0, sticky="NSEW")
        self.username_e = tk.Entry(adduser_window)
        self.username_e.grid(row=0, column=1, sticky="NSEW")
        confirm_b = tk.Button(adduser_window, text='Confirm', command=lambda: self.add_user_confirm())
        confirm_b.grid(row=1, column=0, columnspan=2, sticky="NSEW")
        self.confirm_l = tk.Label(adduser_window)
        self.confirm_l.grid(row=1, column=3, sticky="NSEW")
        close_b = tk.Button(adduser_window, text='Close',
                            command=lambda: adduser_window.destroy() and self.add_user_confirm())
        close_b.grid(row=2, column=0, columnspan=2, sticky="NSEW")

    def add_user_check(self, s):
        x = database.check_operator_number()
        if s == x or self.username_e.get() == '':
            self.confirm_l['text'] = 'Nothing Added!'
            return True
        elif x == s + 1:
            self.confirm_l['text'] = 'Added Successfully!'
            x += 1
            return False

    def add_user_confirm(self):
        x = database.check_operator_number()
        a = self.username_e.get()
        d = datetime.now().strftime('%d-%m-%y %H:%M')
        database.insert_operator(a, d)
        self.add_user_check(x)
        self.cbb_operator.config(values=database.get_operators())

    def list_artefactSN(self, *args):
        self.artefactSN_cbb.config(values=[])
        if self.experiencetype.get() == 'BB150':
            self.artefactSN_cbb.config(values=self.val_artsn[0])
        elif self.experiencetype.get() == 'BB300':
            self.artefactSN_cbb.config(values=self.val_artsn[1])
        elif self.experiencetype.get() == 'BA300':
            self.artefactSN_cbb.config(values=self.val_artsn[2])
        elif self.experiencetype.get() == 'BB500':
            self.artefactSN_cbb.config(values=self.val_artsn[3])

    def copy_to_clipboard(self):
        print(self.experienceid.get())
        if self.experienceid.get() in ' ':
            self.copy_l['text'] = 'Nothing to copy!'
            self.update()
            self.copy_l.after(2000, self.clean_label())
        else:
            pyperclip.copy(str(self.experienceid2.get()))
            self.copy_l['text'] = 'Copied!'
            self.update()
            self.copy_l.after(2000, self.clean_label())

    def clean_label(self):
        self.copy_l['text'] = ''

    def generate_exp_id_random(self):
        dev = str(self.deviceSN.get())
        print(dev)
        if dev == '':
            print('error')
        elif dev[:8] == 'FreeScan':
            dev = dev[8:]
            path = "data.xlsx"
            book = openpyxl.load_workbook(path)
            sheet = book.active
            d = datetime.now().strftime('%d-%m-%y %H:%M')
            max_col = sheet.max_row
            months = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
            now = datetime.now()
            str_mth = now.strftime('%-m')
            month = months[int(str_mth) - 1]
            day = now.strftime('%-d')
            e = dev + month + day + '-' + str(generate_id_random())
            for i in range(1, max_col + 1):
                if e == sheet.cell(row=i, column=1):
                    e = dev + month + day + '-' + str(generate_id_random())
            sheet.append((e, d))
            book.save('data.xlsx')
            print("storing done")
            self.experienceid.insert(0, e)
            self.experienceid.config(state='disabled')
            self.button2.config(state='disabled')

    def insert_to_entry(self, x, y, z, t, r):
        self.deviceSN.insert(0, x)
        self.artefactSN_cbb.set(y)
        self.department.set(z)
        self.operator.set(t)
        # self.experienceid.insert(0, r)

    def get_data_from_xml(self):
        from tkinter import filedialog
        self.filename = filedialog.askopenfilename(initialdir='/Users/kadirakca/git_projects/s3d_db_management/test_db',
                                                   title='Select File',
                                                   filetypes=(('XML files', '*.xml'), ('All files', '*.*')))
        file = xml.dom.minidom.parse(self.filename)
        elements = file.getElementsByTagName('AutoTextCustomFieldValue')
        dev = elements[2].childNodes[0].nodeValue
        art = elements[3].childNodes[0].nodeValue
        dep = elements[0].childNodes[0].nodeValue
        op = elements[1].childNodes[0].nodeValue
        exp = elements[2].childNodes[0].nodeValue
        self.insert_to_entry(dev, art, dep, op, exp)

    def reset(self):
        self.deviceSN.config(state='normal')
        self.operator.set('')
        self.department.set('')
        self.experiencetype.set('')
        self.artefactSN.set('')
        self.certificateSN.config(state='normal')
        self.experienceid.config(state='normal')
        self.deviceSN.delete(0, END)
        self.certificateSN.delete(0, END)
        self.experienceid.delete(0, END)
        self.subject.delete(1.0, END)
        self.button2.config(state='normal')


class PageClassTwo(PageClass):

    def __init__(self, *args, **kwargs):
        PageClass.__init__(self, *args, **kwargs)

        experience_id_ll = tk.Label(self, text="ExperienceID:")
        experience_id_ll.grid(row=1, column=0)

        results_l = tk.Label(self, text='Test Results\n _____________________________', font=fonts.LARGE_FONT)
        results_l.grid(row=2, column=0, columnspan=3, sticky='NSEW')

        self.p2_entry = tk.Label(self, textvariable=x)
        self.p2_entry.grid(row=1, column=1, padx=5, pady=5, ipadx=5, ipady=2, sticky="NSEW")

        copy = tk.Button(self, command=lambda: self.copy_to_clipboard(), text='Copy to Clipboard')
        copy.grid(row=1, column=5, padx=5, pady=5, ipadx=5, ipady=2, sticky="NSEW")

        #self.get_experienceid()

    def get_experienceid(self):
        self.p2_entry['text'] = x
        self.after(1000, self.get_experienceid())


if __name__ == '__main__':
    root = tk.Tk()
    main = DatabaseManagement(root)
    main.pack(side='top', fill='both', expand=True)
    root.wm_title('Database Management Tool for Accuracy Tests')
    root.wm_geometry("650x600")
    root.mainloop()
