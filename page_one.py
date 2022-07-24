import random
import sqlite3
import subprocess
import tkinter as tk
from tkinter import ttk, LEFT, filedialog, END
import xml.dom.minidom
import xlsxwriter
import openpyxl
from datetime import datetime

import page_two

LARGE_FONT = ('Verdana', 18)
NORMAL_FONT = ('Verdana', 12)
SMALL_FONT = ('Verdana', 8)


class PageOne(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.date = None
        self.part_name = None
        self.part_number = None
        self.inspector_name = None
        self.department_name = None
        self.experience_id = None
        self.product_name = None
        self.val_dep = ['Stuttgart', 'HQ', 'HongKong']
        self.val_exptype = ['Not defined 1', 'Not defined 2', 'Not defined 3']
        self.val_artused = ['Ball 150 mm', 'Ball 300 mm', 'Ball 500 mm']
        self.filename = ''
        artefactcalib = None
        self.experienceid = None

        label = tk.Label(self, text='New Test Page', font=LARGE_FONT)
        label.grid(row=0, column=0, pady=10, padx=10, sticky="NSEW", columnspan=3)

        self.deviceSN = tk.Entry(self)
        self.deviceSN.grid(row=1, column=1, padx=5, pady=5, ipadx=5, ipady=2, sticky="NSEW")

        self.operator = tk.Entry(self)
        self.operator.grid(row=2, column=1, padx=5, pady=5, ipadx=5, ipady=2, sticky="NSEW")

        self.department = tk.StringVar()
        self.cbb_department = ttk.Combobox(self, textvariable=self.department)
        self.cbb_department.grid(row=3, column=1, padx=5, pady=5, ipadx=5, ipady=2, sticky="NSEW")
        self.cbb_department['values'] = self.val_dep
        self.cbb_department['state'] = 'readonly'

        self.experiencetype = tk.StringVar()
        self.cbb_experience_type = ttk.Combobox(self, textvariable=self.experiencetype)
        self.cbb_experience_type.grid(row=4, column=1, padx=5, pady=5, ipadx=5, ipady=2, sticky="NSEW")
        self.cbb_experience_type['values'] = self.val_exptype
        self.cbb_experience_type['state'] = 'readonly'

        artefactused = tk.StringVar()
        cbb_artefact_used = ttk.Combobox(self, textvariable=artefactused)
        cbb_artefact_used.grid(row=5, column=1, padx=5, pady=5, ipadx=5, ipady=2, sticky="NSEW")
        cbb_artefact_used['values'] = self.val_artused
        cbb_artefact_used['state'] = 'readonly'

        artefactcalib = tk.Entry(self)
        artefactcalib.grid(row=6, column=1, padx=5, pady=5, ipadx=5, ipady=2, sticky="NSEW")
        # artefactcalib.config(state='disabled')

        subject = tk.Text(self, width=20, height=3)
        subject.grid(row=7, column=1, padx=5, pady=20, ipadx=5, ipady=2, sticky="NSEW")

        self.experienceid = tk.Entry(self)
        self.experienceid.grid(row=9, column=1, padx=5, pady=5, ipadx=5, ipady=2, sticky="NSEW")

        deviceSN_l = tk.Label(self, text="Device SN:")
        deviceSN_l.grid(row=1, column=0)
        operator_l = tk.Label(self, text="Operator")
        operator_l.grid(row=2, column=0)
        department_l = tk.Label(self, text="Department:")
        department_l.grid(row=3, column=0)
        experiencetype = tk.Label(self, text="Experience Type:")
        experiencetype.grid(row=4, column=0)
        artefactused_l = tk.Label(self, text="Artefact Used:")
        artefactused_l.grid(row=5, column=0)
        artefactcalib_l = tk.Label(self, text="Artefact Calibration:")
        artefactcalib_l.grid(row=6, column=0)
        subject_l = tk.Label(self, text="Subject:")
        subject_l.grid(row=7, column=0)
        experienceid_l = tk.Label(self, text="Experience ID:")
        experienceid_l.grid(row=9, column=0)

        button1 = tk.Button(self, text='Fill with XML',
                            command=lambda: self.merge_data(),
                            height=1)
        button1.grid(row=8, column=0, columnspan=2, sticky="NSEW")

        button2 = tk.Button(self, text='Generate ExperienceID',
                            command=lambda: self.generate_exp_id(),
                            height=1)
        button2.grid(row=9, column=3, sticky="NSEW")

        def artefactused_changed(event):
            if artefactused.get() == 'Ball 150 mm':
                artefactcalib.config(state='normal')
                artefactcalib.delete(0, END)
                artefactcalib.insert(0, 'xxxx id_150mm')
                artefactcalib.config(state='disabled')
            elif artefactused.get() == 'Ball 300 mm':
                artefactcalib.config(state='normal')
                artefactcalib.delete(0, END)
                artefactcalib.insert(0, 'xxxx id_300mm')
                artefactcalib.config(state='disabled')
            elif artefactused.get() == 'Ball 500 mm':
                artefactcalib.config(state='normal')
                artefactcalib.delete(0, END)
                artefactcalib.insert(0, 'xxxx id_500mm')
                artefactcalib.config(state='disabled')

        '''def insert_to_entry():
            self.deviceSN.insert(0, get_deviceSN())
            operator.insert(0, get_operator())
            department.set(get_department())
            artefactused.set(get_artefactused())'''

        cbb_artefact_used.bind('<<ComboboxSelected>>', artefactused_changed)

        '''def open_fd():
            self.filename = filedialog.askopenfilename(initialdir='/Users/kadirakca/PycharmProjects',
                                                       title='Select File',
                                                       filetypes=(('XML files', '*.xml'), ('All files', '*.*')))
            return self.filename

        def get_all_cellrtf_in_data(a, b):
            p = a[b].getElementsByTagName('Cell_RTF')
            return p

        def get_rawdata_tag():
            file = xml.dom.minidom.parse(self.filename)
            rowdatas = file.getElementsByTagName('RowData')
            ind = 0
            for rowdata in rowdatas:
                ind = ind + 1
                if rowdata.hasAttribute('RowID') and rowdata.hasAttribute('Row_Type') and rowdata.getAttribute(
                        'RowID') == '0' and rowdata.getAttribute('Row_Type') == '0':
                    cells_rtf = rowdata.getElementsByTagName('Cell_RTF')
                    index_rtf = 0
                    for cell_rtf in cells_rtf:
                        index_rtf = index_rtf + 1
                        if 'Date' in cell_rtf.childNodes[0].nodeValue:
                            ind1 = ind - 1
                            res = get_all_cellrtf_in_data(rowdatas, ind1)
                            return res

        def get_autotext():
            file = xml.dom.minidom.parse(self.filename)
            elements = file.getElementsByTagName('AutoTextCustomFieldValue')
            return elements

        def get_date():
            p = get_rawdata_tag()
            self.date = p[1].childNodes[0].nodeValue[411:-7]
            return self.date

        def get_deviceSN():
            elements = get_autotext()
            self.product_name = elements[2].childNodes[0].nodeValue
            return self.product_name

        def get_artefactused():
            elements = get_autotext()
            self.part_name = elements[3].childNodes[0].nodeValue
            return self.part_name

        def get_department():
            elements = get_autotext()
            self.department_name = elements[0].childNodes[0].nodeValue
            return self.department_name

        def get_operator():
            elements = get_autotext()
            self.inspector_name = elements[1].childNodes[0].nodeValue
            return self.inspector_name

        def get_partnumber():
            elements = get_autotext()
            self.part_number = elements[4].childNodes[0].nodeValue
            return self.part_number

        def get_experienceid():
            elements = get_autotext()
            self.experience_id = elements[2].childNodes[0].nodeValue
            return self.experience_id

        def merge_data():
            open_fd()
            data = [get_deviceSN(), get_experienceid(), get_partnumber()]
            insert_to_entry()
            self.experienceid.config(state='disabled')
            return data'''

    def generate_exp_id(self):
        wb = openpyxl.load_workbook('exp_id.xlsx')
        sh = wb.active
        months = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
        now = datetime.now()
        str_mth = now.strftime('%-m')
        month = months[int(str_mth) - 1]
        day = now.strftime('%-d')
        x = sh.max_row
        last_cell = sh.cell(row=x, column=1)
        if last_cell.value is None:
            experience_id = str(self.deviceSN.get()) + month + day + '-000'
            sh['A1'] = str(experience_id)
            self.experienceid.insert(0, experience_id)
        else:
            last_no = last_cell.value[-3:]
            experience_id = str(self.deviceSN.get()) + month + day + str(int(last_no) + 1)
            sh.cell(row=sh.max_row + 1, column=1)
            self.experienceid.insert(0, experience_id)
            print(experience_id)
            page_two.PageTwo.insert_experience_id_l('dsada')

    def insert_to_entry(self):
        self.deviceSN.insert(0, self.get_deviceSN())
        self.operator.insert(0, self.get_operator())
        self.department.set(self.get_department())
        # artefactused.set(self.get_artefactused())

    def read_XML(self):

        pass

    def open_fd(self):
        self.filename = filedialog.askopenfilename(initialdir='/Users/kadirakca/PycharmProjects',
                                                   title='Select File',
                                                   filetypes=(('XML files', '*.xml'), ('All files', '*.*')))
        return self.filename

    def get_all_cellrtf_in_data(self, a, b):
        p = a[b].getElementsByTagName('Cell_RTF')
        return p

    def get_rawdata_tag(self):
        file = xml.dom.minidom.parse(self.filename)
        rowdatas = file.getElementsByTagName('RowData')
        ind = 0
        for rowdata in rowdatas:
            ind = ind + 1
            if rowdata.hasAttribute('RowID') and rowdata.hasAttribute('Row_Type') and rowdata.getAttribute(
                    'RowID') == '0' and rowdata.getAttribute('Row_Type') == '0':
                cells_rtf = rowdata.getElementsByTagName('Cell_RTF')
                index_rtf = 0
                for cell_rtf in cells_rtf:
                    index_rtf = index_rtf + 1
                    if 'Date' in cell_rtf.childNodes[0].nodeValue:
                        ind1 = ind - 1
                        res = self.get_all_cellrtf_in_data(rowdatas, ind1)
                        return res

    def get_autotext(self):
        file = xml.dom.minidom.parse(self.filename)
        elements = file.getElementsByTagName('AutoTextCustomFieldValue')
        return elements

    def get_date(self):
        p = self.get_rawdata_tag()
        self.date = p[1].childNodes[0].nodeValue[411:-7]
        return self.date

    def get_deviceSN(self):
        elements = self.get_autotext()
        self.product_name = elements[2].childNodes[0].nodeValue
        return self.product_name

    def get_artefactused(self):
        elements = self.get_autotext()
        self.part_name = elements[3].childNodes[0].nodeValue
        return self.part_name

    def get_department(self):
        elements = self.get_autotext()
        self.department_name = elements[0].childNodes[0].nodeValue
        return self.department_name

    def get_operator(self):
        elements = self.get_autotext()
        self.inspector_name = elements[1].childNodes[0].nodeValue
        return self.inspector_name

    def get_partnumber(self):
        elements = self.get_autotext()
        self.part_number = elements[4].childNodes[0].nodeValue
        return self.part_number

    def get_experienceid(self):
        elements = self.get_autotext()
        self.experience_id = elements[2].childNodes[0].nodeValue
        return self.experience_id

    def merge_data(self):
        self.open_fd()
        data = [self.get_deviceSN(), self.get_experienceid(), self.get_partnumber()]
        self.insert_to_entry()
        self.experienceid.config(state='disabled')
        return data
